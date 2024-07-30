from datetime import datetime
from time import time
from warnings import simplefilter
from cssutils import log  # se importa para descativar los WARNINGS al importar la tabla html usando  tablepyxl
import logging


from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.datetime import from_excel

from tablepyxl import tablepyxl

from dbconfig import save_to_db


def importa_informe(archivo_cz, columnas_cz, db_name, db_table, first_data_row, col_name_row, max_col, hoja, col_check_flag):
    time_start = time()

    # Verifica si la ruta contiene el texto "web_" para así primero pasar el archivo html (exportado de la web de monitoreos) a wb usado por openpyxl
    # si no contiene dicho texto entonces se carga el archivo como un xlsx normal
    if str(archivo_cz).find('APP MONITOREO_') == -1 and str(archivo_cz).find('BACKOFFICE_BO') == -1:
        simplefilter("ignore")  # Desactiva el warning de graficos cuando se utiliza load_workbook
        wb = load_workbook(filename=archivo_cz, data_only=True, keep_vba=False, keep_links=False, read_only=True)
    else:
        log.setLevel(logging.CRITICAL)  # Suprime los WARNINGS de CSSUTILS, usado por tablepyxl
        doc = open(archivo_cz, 'r', encoding='latin-1')
        tbl = doc.read()
        wb = tablepyxl.document_to_workbook(tbl)
        doc.close()

    if hoja not in wb.sheetnames:
        print(f"No se encontro la hoja '{hoja}'")
        quit()

    ws = wb[hoja]
    # print(ws.cell(row=5, column=1).value)

    # Verifica si las cabeceras del XL corresponden a las registradas en el archivo de configuración
    for row in ws.iter_rows(min_row=col_name_row, max_row=col_name_row, max_col=max_col, values_only=True):
        # if row[column_index_from_string(col_check_flag) - 1] is None:  # iterrows empieza desde 0 por eso se resta 1 cuando se combierte la letra de la col a número
        #     break
        for col in columnas_cz:
            valor = row[column_index_from_string(col['col']) - 1]
            if str(valor).upper().strip() != str(col['colName']).upper().strip():
                print(f'ERROR en cabecera: py: "{col["colName"]}" - xl: "{valor}" - col: "{col["col"]}" - archivo: {archivo_cz}')
                quit()
            # else:
            #     print(f'{valor}...OK')

    # Empieza con el proceso de importación leyendo la data del archivo excel
    lista_evas = []
    lista_sn = []
    for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row, max_col=max_col, values_only=True):
        lista_temp = {}
        if row[column_index_from_string(col_check_flag) - 1] is None:  # iterrows empieza desde 0 por eso se resta 1 cuando se combierte la letra de la col a número
            break

        for col in columnas_cz:
            valor = row[column_index_from_string(col['col']) - 1]
            if col['tipo'] == 'str':
                valor = str(valor)
            elif col['tipo'] == 'cf':
                try:
                    valor = valor / 100.0
                except TypeError as t:
                    print(f'{t} - col: {col["col"]} - valor: {valor}')
                    raise
            elif col['tipo'] == 'hora':
                valor = valor.strftime("%H:%M:%S.%f")
            elif col['tipo'] == 'trim_str':
                if valor is not None:
                    valor = str(valor).strip()
            elif col['tipo'] == 'xl_date':
                if valor is not None:
                    if valor < 4000:
                        valor = None
                    else:
                        valor = from_excel(valor)
            elif col['tipo'] == 'str_to_datetime':
                if valor is not None:
                    valor = datetime.strptime(valor, '%d/%m/%Y %H:%M:%S')
            elif col['tipo'] == 'str_to_datetime_webevas':
                if valor is not None:
                    valor = datetime.strptime(valor.replace('p.m.', 'pm').replace('a.m.', 'am'), '%d/%m/%Y %I:%M:%S %p')
            elif col['tipo'] == 'str_to_int':
                if valor is not None:
                    valor = int(str(valor).strip())
            elif col['tipo'] == 'no_duplicado':
                while valor in lista_sn:
                    valor = str(valor) + "'"
                lista_sn.append(valor)

            if isinstance(valor, str):  # Si valor es string le hcace strip antes de agregarlo a la lista temporal
                valor = valor.strip()
            lista_temp[col['sqlName']] = valor
        lista_evas.append(lista_temp)

    ws = None
    wb.close()
    save_to_db(db_name, db_table, lista_evas)
    simplefilter("default")  # Habilita los warnings nuevamente
    time_end = time()
    print(f"TIEMPO: {time_end - time_start}")
